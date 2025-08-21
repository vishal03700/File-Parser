import pandas as pd
import PyPDF2
import pdfplumber
import io
import json
import logging
from typing import Dict, Any, List
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class FileParser:
    """File parser for different file types."""
    
    @staticmethod
    def parse_csv(file_content: bytes) -> Dict[str, Any]:
        """Parse CSV file content."""
        try:
            df = pd.read_csv(io.BytesIO(file_content))
            
            # Convert DataFrame to dictionary
            data = {
                'headers': df.columns.tolist(),
                'rows': df.to_dict('records'),
                'total_rows': len(df),
                'columns': len(df.columns)
            }
            
            return {
                'success': True,
                'data': data,
                'content_type': 'csv'
            }
        except Exception as e:
            logger.error(f"Error parsing CSV: {str(e)}")
            return {
                'success': False,
                'error': f"Failed to parse CSV: {str(e)}",
                'content_type': 'csv'
            }
    
    @staticmethod
    def parse_excel(file_content: bytes) -> Dict[str, Any]:
        """Parse Excel file content."""
        try:
            workbook = load_workbook(io.BytesIO(file_content))
            sheets_data = {}
            total_rows = 0
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Convert sheet to list of lists
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(list(row))
                
                if data:
                    headers = data[0] if data else []
                    rows = data[1:] if len(data) > 1 else []
                    
                    sheets_data[sheet_name] = {
                        'headers': headers,
                        'rows': [dict(zip(headers, row)) for row in rows],
                        'total_rows': len(rows)
                    }
                    total_rows += len(rows)
            
            return {
                'success': True,
                'data': {
                    'sheets': sheets_data,
                    'sheet_names': list(sheets_data.keys()),
                    'total_rows': total_rows
                },
                'content_type': 'excel'
            }
        except Exception as e:
            logger.error(f"Error parsing Excel: {str(e)}")
            return {
                'success': False,
                'error': f"Failed to parse Excel: {str(e)}",
                'content_type': 'excel'
            }
    
    @staticmethod
    def parse_pdf(file_content: bytes) -> Dict[str, Any]:
        """Parse PDF file content."""
        try:
            text_content = []
            
            # Try with pdfplumber first (better for text extraction)
            with pdfplumber.open(io.BytesIO(file_content)) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    text = page.extract_text()
                    if text:
                        text_content.append({
                            'page': page_num,
                            'content': text.strip()
                        })
            
            # Fallback to PyPDF2 if pdfplumber fails
            if not text_content:
                pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
                for page_num, page in enumerate(pdf_reader.pages, 1):
                    text = page.extract_text()
                    if text:
                        text_content.append({
                            'page': page_num,
                            'content': text.strip()
                        })
            
            return {
                'success': True,
                'data': {
                    'pages': text_content,
                    'total_pages': len(text_content),
                    'full_text': '\n\n'.join([page['content'] for page in text_content])
                },
                'content_type': 'pdf'
            }
        except Exception as e:
            logger.error(f"Error parsing PDF: {str(e)}")
            return {
                'success': False,
                'error': f"Failed to parse PDF: {str(e)}",
                'content_type': 'pdf'
            }
    
    @classmethod
    def parse_file(cls, file_content: bytes, file_type: str, filename: str) -> Dict[str, Any]:
        """Parse file based on its type."""
        filename_lower = filename.lower()
        
        if filename_lower.endswith('.csv') or 'csv' in file_type:
            return cls.parse_csv(file_content)
        elif filename_lower.endswith(('.xlsx', '.xls')) or 'excel' in file_type or 'spreadsheet' in file_type:
            return cls.parse_excel(file_content)
        elif filename_lower.endswith('.pdf') or 'pdf' in file_type:
            return cls.parse_pdf(file_content)
        else:
            return {
                'success': False,
                'error': f"Unsupported file type: {file_type}. Supported types: CSV, Excel (.xlsx, .xls), PDF",
                'content_type': 'unknown'
            }